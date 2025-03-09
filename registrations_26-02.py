# -*- coding: utf-8 -*-
"""
Created on Thu Feb 27 02:04:29 2025

@author: Artem
"""

import json
import logging
import argparse
from typing import Dict, Tuple, List, Optional
from urllib.parse import parse_qs, unquote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Константы
class Config:
    CITY = "NSK"
    PHONE_PREFIX = "7"
    MIN_BIRTH_YEAR = 1950
    MAX_BIRTH_YEAR = 2009
    EXPECTED_PHONE_LENGTH = 11
    NUMBER_OF_SEMIFINAL_PARTICIPANTS = 55
    DATE_RANGE_START = pd.to_datetime("2025-01-01")
    DATE_RANGE_END = pd.to_datetime("2025-01-31")
    
    DELETED_STATUS = "удалена"
    CLASSIC_GAME = "[классика]"
    NOVICE_GAME = "[новички]"
    UTM_COLUMNS = ["source", "medium", "campaign", "content", "term"]
    COMMENT_COLUMN = "Комментарий"
    STATUS_COLUMN = "Статус команды"
    ARRIVED_COLUMN = "Пришло людей"
    PARTIAL_DUPLICATE_COLUMNS = ["Имя", "Email", "Телефон"]   
    
    INVALID_PHONES = ["77777777777", "79999999999", "71234567890", "79123456789"]
    SOURCE_KEYWORDS = {
        "app": [r"\bΜ$"],
        "admin": [r"\bАдмин\b"],
        "web": [r"\bИнтернет\b"],
        "certificate": [r"\bсертификат\b"],
        "sarafan": ["посоветовали", "друзья", "узнал от коллег", "от друзей", "от знакомых", "от коллег", "от подруги", "от друга", "сарафан"],
        "ads": ["реклам", "таргет"],
        "sxodim": ["sxodim", "сходим"],
        "instagram": ["инстаграм", "instagram", "инсты", "инстаграмм"]
    }

# Настройка логирования
def setup_logging(log_level="INFO"):
    logging.basicConfig(
        level=log_level.upper(),
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler("data_processing.log"),
            logging.StreamHandler()
        ]
    )

# Загрузка словарей
def load_dictionaries() -> Tuple[Dict, Dict, Dict]:
    """
    Загружает словари имён и доменов из JSON-файлов.
    
    Returns:
        tuple: (name_variants, domain_corrections)
    """
    try:
        with open("name_variants.json", "r", encoding="utf-8") as f:
            name_variants = json.load(f)
        with open("domain_corrections.json", "r", encoding="utf-8") as f:
            domain_corrections = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"Ошибка при загрузке словарей: {type(e).__name__} - {e}")
        raise
    reverse_name_mapping = {}
    for full_name, variants in name_variants.items():
        for variant in variants:
            reverse_name_mapping[variant.lower()] = full_name

    return name_variants, domain_corrections, reverse_name_mapping

# Вспомогательные функции обработки текста
def normalize_name(name: str, reverse_name_mapping: Dict) -> str:
    """
    Нормализует имя, удаляя фамилию и приводя имя к нужному формату.

    Args:
        name (str): Имя для обработки (может содержать фамилию).
        reverse_name_mapping (dict): Словарь для поиска полного имени.

    Returns:
        str: Нормализованное имя или исходное значение, если входные данные не строка.
    """    
    if not isinstance(name, str):
        return name
    # Разделяем строку на слова (имя и фамилия)
    words = name.strip().split()
    if not words:  # Если строка пуста после разделения
        return name
    # if len(words) > 1:
        # logging.info(f"Обнаружена строка с возможной фамилией: {name}")
    # Ищем совпадения в словаре
    for word in words:
        lower_word = word.lower()
        if lower_word in reverse_name_mapping:
            return reverse_name_mapping[lower_word].capitalize()
    
    return words[0].capitalize()

def repair_email_domain(email: str, domain_corrections: Dict) -> str:
    """
    Исправляет ошибки в доменах email.
    
    Args:
        email (str): Email для обработки.
        domain_corrections (dict): Словарь исправлений доменов.
    
    Returns:
        str: Исправленный email или исходное значение.
    """
    if not isinstance(email, str):
        return email
    email = email.replace(" ", "").lower()
    if "@" not in email:
        logging.warning(f"Некорректный email: {email}")
        return email
    local_part, domain = email.split("@")
    for correct_domain, incorrect_domains in domain_corrections.items():
        if domain in incorrect_domains:
            return f"{local_part}@{correct_domain}"
    return email

def decode_utm_string(utm_string: str) -> str:
    if not isinstance(utm_string, str):
        return utm_string
    replacements = {
        "%3D": "=", "%253D": "=", "%26": "&",
        "%2F": "/", "%3F": "?"
    }
    for old, new in replacements.items():
        utm_string = utm_string.replace(old, new)
    return unquote(utm_string)

def parse_utm(utm_string: str) -> pd.Series:
    """
    Извлекает UTM-метки из строки, игнорируя другие параметры до UTM-меток.
    
    Args:
        utm_string (str): Строка с UTM-метками.
    
    Returns:
        pd.Series: Серия с извлечёнными UTM-метками.
    """
    utm_params = {"source": "", "medium": "", "campaign": "", "content": "", "term": ""}
    if not isinstance(utm_string, str) or "utm_" not in utm_string:
        return pd.Series(utm_params)
    try:
        utm_string = decode_utm_string(utm_string)
        utm_start = utm_string.find("utm_")
        if utm_start == -1:
            return pd.Series(utm_params)
        utm_part = utm_string[utm_start:]
        params = parse_qs(utm_part, keep_blank_values=True)
        for key, value in params.items():
            if key.startswith("utm_"):
                param_name = key.replace("utm_", "")
                if param_name in utm_params:
                    utm_params[param_name] = value[0] if value else ""
    except Exception as e:
        logging.warning(f"Ошибка при парсинге UTM-меток: {utm_string} - {type(e).__name__} - {e}")
    return pd.Series(utm_params)

# Функции очистки данных
def drop_columns(df: pd.DataFrame) -> pd.DataFrame:
    columns_to_drop = ["Город", "Запись", "Кастомные поля"]
    df.drop(columns=columns_to_drop, errors="ignore", inplace=True)
    return df

def rename_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={
        "Заголовок": "Игра",
        "Название": "Номер пакета",
        "Дата и время": "Дата игры",
        "Капитан": "Имя",
        "Дата": "Дата и время регистрации",
        "Кол-во": "Планировалось людей"
    })

def filter_dates(df: pd.DataFrame) -> pd.DataFrame:
    df["Дата игры"] = pd.to_datetime(df["Дата игры"], format='%d.%m.%Y %H:%M', errors='coerce')
    df = df[(df["Дата игры"] >= Config.DATE_RANGE_START) & (df["Дата игры"] <= Config.DATE_RANGE_END)]
    
    df["Дата игры"] = df["Дата игры"].dt.strftime("%d.%m.%Y")
    logging.info(f"Количество строк после фильтрации по датам: {len(df)}")
    
    return df

def clean_phone_numbers(df: pd.DataFrame) -> pd.DataFrame:
    df["Телефон"] = (df["Телефон"].fillna("").astype(str)
                     .str.replace(r"[^0-9]", "", regex=True)
                     .str.replace(r"^8", Config.PHONE_PREFIX, regex=True))
    initial_len = len(df)
    df = df[~df["Телефон"].isin(Config.INVALID_PHONES)]
    removed = initial_len - len(df)
    if removed > 0:
        logging.info(f"Удалено строк с некорректными телефонами: {removed}")
    return df

def clean_text_columns(df: pd.DataFrame, reverse_name_mapping: Dict) -> pd.DataFrame:
    text_columns = ["Имя", "Команда", "Комментарий"]
    for col in text_columns:
        df[col] = df[col].fillna("").astype(str).str.replace(r"  ", " ", regex=True).str.replace("ё", "е").str.replace("Ё", "Е")
    df["Имя"] = df["Имя"].map(lambda x: normalize_name(x, reverse_name_mapping))
    return df

def clean_comments(df: pd.DataFrame) -> pd.DataFrame:
    mask_matching_email = (df["Комментарий"] == df["Email"]) & (df["Email"] != "")
    cleaned_comments = mask_matching_email.sum()
    if cleaned_comments > 0:
        df.loc[mask_matching_email, "Комментарий"] = ""
        logging.info(f"Очищено комментариев, совпадающих с email: {cleaned_comments}")
    return df

def process_utm(df: pd.DataFrame) -> pd.DataFrame:
    utm_columns = df["Данные utm"].apply(parse_utm)
    df = pd.concat([df, utm_columns], axis=1)
    return df.drop(columns=["Данные utm"], errors="ignore")

def determine_source(df: pd.DataFrame) -> pd.DataFrame:
    df["Комментарий"] = df["Комментарий"].fillna("")
    # Устанавливаем source = "admin" для строк с "отборочная" или "полуфинал" в столбце "Игра"
    df.loc[df["Игра"].str.contains(r"отборочная|полуфинал", case=False, na=False), "source"] = "admin"
    # Остальная логика определения источника
    for source, patterns in Config.SOURCE_KEYWORDS.items():
        for pattern in patterns:
            if source == "app":
                df.loc[df["Команда"].str.endswith("Μ"), "source"] = "app"
            elif source == "admin":
                df.loc[df["Имя"].str.contains(pattern, case=False, na=False), "source"] = "admin"
            else:
                df.loc[df["Комментарий"].str.contains(pattern, case=False, na=False), "source"] = source
    return df

def filter_semifinal_participants(df: pd.DataFrame) -> pd.DataFrame:
    # Находим строки с "отборочная" или "полуфинал" в столбце "Игра"
    semifinal_mask = df["Игра"].str.contains(r"отборочная|полуфинал", case=False, na=False)
    semifinal_df = df[semifinal_mask]
    
    # Если количество таких строк больше заданного лимита
    if len(semifinal_df) > Config.NUMBER_OF_SEMIFINAL_PARTICIPANTS:
        # Находим строки с пустым значением в "пришло людей"
        empty_arrived_mask = semifinal_df["Пришло людей"].isna() | (semifinal_df["Пришло людей"] == "")
        empty_arrived_df = semifinal_df[empty_arrived_mask]
        
        # Если есть строки с пустым "пришло людей"
        if not empty_arrived_df.empty:
            # Вычисляем, сколько строк нужно удалить
            excess_count = len(semifinal_df) - Config.NUMBER_OF_SEMIFINAL_PARTICIPANTS
            rows_to_remove = min(excess_count, len(empty_arrived_df))
            
            # Удаляем случайные строки с пустым "пришло людей"
            if rows_to_remove > 0:
                rows_to_drop = empty_arrived_df.index[:rows_to_remove]
                df = df.drop(rows_to_drop)
                logging.info(f"Удалено строк с пустым 'пришло людей' для отборочных/полуфинальных игр: {rows_to_remove}")
    
    return df

def filter_final_and_ceremony(df: pd.DataFrame) -> pd.DataFrame:
    initial_len = len(df)
    
    # Фильтруем строки, где:
    # 1) Есть "ОТКРЫТИЕ" или "Церемония"
    # 2) Но при этом нет "отборочная" или "полуфинал" (для строк с "ОТКРЫТИЕ")
    mask_opening = df["Игра"].str.contains(r"ОТКРЫТИЕ", case=False, na=False)
    mask_ceremony = df["Игра"].str.contains(r"Церемония", case=False, na=False)
    mask_qualifying = df["Игра"].str.contains(r"отборочная|полуфинал", case=False, na=False)
    
    # Удаляем строки, где:
    # - Есть "Церемония" ИЛИ
    # - Есть "ОТКРЫТИЕ", но нет "отборочная" или "полуфинал"
    mask_to_remove = mask_ceremony | (mask_opening & ~mask_qualifying)
    df = df[~mask_to_remove]
    
    removed = initial_len - len(df)
    if removed > 0:
        logging.info(f"Удалено строк с 'открытие' или 'церемония' в столбце 'Игра': {removed}")
    return df

def rename_games(df: pd.DataFrame) -> pd.DataFrame:
    """
    Обрабатывает значения в столбце 'Игра' согласно заданным правилам:
    - Заменяет 'Квиз, плиз! {CITY}' на '[классика]'
    - Удаляет упоминания 'Квиз, плиз!' и '{CITY}'
    - Исправляет оставшиеся '.jpeg' на '[.jpeg]'

    Args:
        df (pd.DataFrame): DataFrame с данными.

    Returns:
        pd.DataFrame: DataFrame с обработанным столбцом 'Игра'.
    """
    if "Игра" not in df.columns:
        logging.warning("Столбец 'Игра' отсутствует в данных")
        return df

    # Шаг 1: Заменяем "Квиз, плиз! {CITY}" на "[классика]"
    classic_pattern = f"Квиз, плиз! {Config.CITY}"
    df["Игра"] = df["Игра"].replace(classic_pattern, "[классика]")

    # Шаг 2: Удаляем "Квиз, плиз!" и "{CITY}" из всех значений
    df["Игра"] = df["Игра"].str.replace("Квиз, плиз!", "", regex=False)
    df["Игра"] = df["Игра"].str.replace(Config.CITY, "", regex=False)

    # Шаг 3: Исправляем оставшиеся ".jpeg" на "[.jpeg]"
    df["Игра"] = df["Игра"].str.replace(".jpeg", "[.jpeg]", regex=False)

    # Удаляем лишние пробелы, которые могли появиться после замен
    df["Игра"] = df["Игра"].str.strip()

    logging.info("Столбец 'Игра' успешно обработан")
    return df

# Функции удаления дубликатов
def remove_duplicates(
    df: pd.DataFrame,
    key_columns: List[str],
    scenario: str,
    additional_conditions: Optional[Dict] = None
) -> pd.DataFrame:
    """
    Удаляет дубликаты в DataFrame на основе ключевых столбцов и сценария.

    Args:
        df (pd.DataFrame): DataFrame для обработки.
        key_columns (List[str]): Столбцы для определения дубликатов.
        scenario (str): Сценарий удаления дубликатов ("same_game_package" или "same_date").
        additional_conditions (Optional[Dict]): Дополнительные условия для сценария.

    Returns:
        pd.DataFrame: DataFrame без дубликатов.
    """
    logging.info(f"Начало удаления дубликатов для сценария: {scenario}")

    # Проверяем наличие всех необходимых столбцов
    required_columns = key_columns + Config.UTM_COLUMNS + [
        Config.COMMENT_COLUMN, Config.ARRIVED_COLUMN, Config.STATUS_COLUMN
    ]
    if "Игра" in df.columns and scenario == "same_game_package":
        required_columns.append("Игра")
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.warning(f"Отсутствуют столбцы: {missing_columns}. Удаление дубликатов невозможно.")
        return df

    # Создаём копию DataFrame для безопасной работы
    df = df.copy()

    # Специальная обработка для сценария "same_game_package"
    if scenario == "same_game_package":
        # Фильтруем строки для "[классика]" и "[новички]"
        game_mask = df["Игра"].isin([Config.CLASSIC_GAME, Config.NOVICE_GAME])
        target_df = df[game_mask].copy()
        non_target_df = df[~game_mask].copy()

        # Находим дубликаты по ключевым столбцам для обычных игр (не "[классика]" и не "[новички]")
        duplicated_mask_non_target = non_target_df.duplicated(subset=key_columns, keep=False)
        # Находим дубликаты по "Номер пакета", "Имя", "Email", "Телефон" для "[классика]" и "[новички]"
        package_key_columns = [col for col in key_columns if col != "Игра"]
        duplicated_mask_target = target_df.duplicated(subset=package_key_columns, keep=False)

        # Разделяем дубликаты и недубликаты для каждой группы
        duplicates_non_target_df = non_target_df[duplicated_mask_non_target].copy()
        non_duplicates_non_target_df = non_target_df[~duplicated_mask_non_target].copy()
        duplicates_target_df = target_df[duplicated_mask_target].copy()
        non_duplicates_target_df = target_df[~duplicated_mask_target].copy()

        # Объединяем дубликаты для обработки
        duplicates_df = pd.concat([duplicates_non_target_df, duplicates_target_df]).drop_duplicates()
        non_duplicates_df = pd.concat([non_duplicates_non_target_df, non_duplicates_target_df], ignore_index=True)

        # Группируем дубликаты
        grouped = duplicates_df.groupby(
            package_key_columns if duplicates_df["Игра"].isin([Config.CLASSIC_GAME, Config.NOVICE_GAME]).any()
            else key_columns,
            as_index=False
        )
    elif scenario == "same_date":
        # Находим дубликаты по ключевым столбцам
        duplicated_mask = df.duplicated(subset=key_columns, keep=False)
        if not duplicated_mask.any():
            logging.info("Дубликаты не найдены")
            return df
        duplicates_df = df[duplicated_mask].copy()
        non_duplicates_df = df[~duplicated_mask].copy()
        grouped = duplicates_df.groupby(key_columns, as_index=False)
    else:
        raise ValueError(f"Неподдерживаемый сценарий: {scenario}")

    # Список для хранения итоговых строк
    final_rows = []

    for _, group in grouped:
        # Проверяем значения в столбце "Пришло людей"
        arrived_notna = group[Config.ARRIVED_COLUMN].notna() & (group[Config.ARRIVED_COLUMN] != "")

        if arrived_notna.all():
            # Если во всех строках "Пришло людей" непустое, сохраняем все строки
            final_rows.append(group)
            continue

        if arrived_notna.any():
            # Сохраняем только строки с непустым "Пришло людей"
            rows_to_keep = group[arrived_notna]
            final_rows.append(rows_to_keep)
            removed_rows = group[~arrived_notna]
            
            for _, row in removed_rows.iterrows():
                logging.debug(f"Удалена строка: {row[key_columns].to_dict()}")
            continue

        # Если все строки с пустым "Пришло людей"
        # Сортируем группу, чтобы записи со статусом "удалена" были первыми
        group = group.sort_values(
            by=Config.STATUS_COLUMN,
            key=lambda x: x == Config.DELETED_STATUS,
            ascending=False
        )

        # Выбираем первую строку как базовую
        base_row = group.iloc[0].copy()

        # Объединяем UTM-метки (берём первые непустые значения)
        for utm_col in Config.UTM_COLUMNS:
            utm_values = group[utm_col].dropna()
            if utm_values.empty and base_row[utm_col] == "":
                continue
            base_row[utm_col] = utm_values.iloc[0] if not utm_values.empty else base_row[utm_col]

        # Объединяем комментарии (берём первый непустой или конкатенируем через пробел)
        comments = group[Config.COMMENT_COLUMN].dropna().unique()
        base_row[Config.COMMENT_COLUMN] = " ".join(comments) if comments.size > 0 else base_row[Config.COMMENT_COLUMN]

        final_rows.append(base_row.to_frame().T)
        removed_rows = group.iloc[1:]
        
        for _, row in removed_rows.iterrows():
            logging.debug(f"Удалена строка: {row[key_columns].to_dict()}")

    # Объединяем результаты
    if final_rows:
        final_duplicates_df = pd.concat(final_rows, ignore_index=True)
        df = pd.concat([non_duplicates_df, final_duplicates_df], ignore_index=True)
    else:
        df = non_duplicates_df

    logging.info(f"Удаление дубликатов завершено для сценария {scenario}. Строк после удаления: {len(df)}")
    return df

def highlight_partial_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Находит потенциальные дубликаты, где у хотя бы одного дубликата пустое 'Пришло людей'
    и есть различия в полях 'Команда', 'Имя', 'Email' или 'Телефон'. Подсвечивает различающиеся
    значения жёлтым цветом в итоговом Excel-файле.

    Args:
        df (pd.DataFrame): DataFrame с данными.

    Returns:
        pd.DataFrame: DataFrame с добавленной колонкой '_highlight_info' для подсветки различий.
    """
    logging.info("Начало поиска потенциальных дубликатов для подсветки")

    # Столбцы для проверки дубликатов (аналогично remove_duplicates)
    key_columns_game = ["Игра", "Номер пакета", "Имя", "Email", "Телефон"]
    # Столбцы для проверки дубликатов (аналогично remove_same_day_no_show_duplicates)
    key_columns_day = ["Дата игры", "Команда", "Имя", "Email", "Телефон"]
    # Поля для проверки различий
    diff_columns = ["Команда", "Имя", "Email", "Телефон"]
    # Столбцы с UTM-метками (для логирования, если нужно)
    utm_columns = ["source", "medium", "campaign", "content", "term"]

    # Проверяем наличие всех необходимых столбцов
    required_columns = set(key_columns_game + key_columns_day + utm_columns + ["Пришло людей"])
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.warning(f"Отсутствуют столбцы: {missing_columns}. Подсветка потенциальных дубликатов невозможна.")
        return df

    # Создаём копию DataFrame для безопасной работы
    df = df.copy()

    # Создаём временный столбец для нормализованной команды (без '*' и 'Μ' в конце)
    df["Команда_норм"] = df["Команда"].str.replace(r"[\*Μ]$", "", regex=True)

    # Создаём словарь для хранения информации о подсветке
    df["_highlight_info"] = [[] for _ in range(len(df))]

    # Функция для проверки различий в группе
    def check_differences(group, key_columns, is_day_check=False):
        if len(group) < 2:
            return

        # Проверяем, есть ли хотя бы один дубликат с пустым "Пришло людей"
        arrived_notna = group["Пришло людей"].notna() & (group["Пришло людей"] != "")
        if arrived_notna.all():
            logging.debug(f"Группа пропущена (все 'Пришло людей' непустые): {group[key_columns].iloc[0].to_dict()}")
            return  # Все "Пришло людей" непустые, пропускаем

        # Проверяем различия в diff_columns
        for col in diff_columns:
            if col == "Команда" and is_day_check:
                values = group["Команда_норм"].values
            else:
                values = group[col].values
            if not all(v == values[0] for v in values):
                # Нашли различие, добавляем информацию о подсветке
                for idx in group.index:
                    if col == "Команда" and is_day_check:
                        df.at[idx, "_highlight_info"].append("Команда")
                    else:
                        df.at[idx, "_highlight_info"].append(col)
                logging.debug(f"Найдено различие в столбце '{col}' для группы: {group[key_columns].to_dict()}")

    # Функция для поиска "почти дубликатов"
    def find_almost_duplicates(key_columns, is_day_check=False):
        # Для каждого столбца из diff_columns создаём комбинацию ключей без этого столбца
        for diff_col in diff_columns:
            if is_day_check and diff_col == "Команда":
                check_columns = [col if col != "Команда" else "Команда_норм" for col in key_columns]
            else:
                check_columns = key_columns.copy()
            if diff_col in check_columns:
                check_columns.remove(diff_col)
            else:
                if is_day_check and diff_col == "Команда":
                    check_columns.remove("Команда_норм")
                else:
                    continue

            logging.debug(f"Проверка почти дубликатов без столбца '{diff_col}': {check_columns}")
            duplicated_mask = df.duplicated(subset=check_columns, keep=False)
            if duplicated_mask.any():
                duplicates_df = df[duplicated_mask].copy()
                grouped = duplicates_df.groupby(check_columns, as_index=False)
                for _, group in grouped:
                    check_differences(group, key_columns, is_day_check)

    # Проверяем дубликаты по логике remove_duplicates (одна игра и номер пакета)
    logging.info("Проверка почти дубликатов по логике remove_same_game_duplicates")
    find_almost_duplicates(key_columns_game, is_day_check=False)

    # Проверяем дубликаты по логике remove_same_day_no_show_duplicates (разные игры в один день)
    key_columns_day_normalized = ["Дата игры", "Команда_норм", "Имя", "Email", "Телефон"]
    logging.info("Проверка почти дубликатов по логике remove_same_day_duplicates")
    find_almost_duplicates(key_columns_day_normalized, is_day_check=True)

    # Удаляем временный столбец
    df = df.drop(columns=["Команда_норм"])

    # Логируем строки с подсветкой
    highlighted_rows = df[df["_highlight_info"].apply(len) > 0]
    if not highlighted_rows.empty:
        logging.info(f"Найдено строк для подсветки: {len(highlighted_rows)}")
        for idx, row in highlighted_rows.iterrows():
            logging.debug(f"Строка {idx} для подсветки: {row['_highlight_info']}, данные: {row[key_columns_day].to_dict()}")

    logging.info("Поиск потенциальных дубликатов для подсветки завершён")
    return df

def clean_data(df: pd.DataFrame, reverse_name_mapping: Dict, domain_corrections: Dict) -> pd.DataFrame:
    """
    Очищает и обрабатывает данные.
    
    Args:
        df (pd.DataFrame): DataFrame для очистки.
        reverse_name_mapping (dict): Словарь для поиска полного имени.
        domain_corrections (dict): Словарь исправлений доменов.
    
    Returns:
        pd.DataFrame: Очищенный DataFrame.
    """
    df = drop_columns(df)
    df = rename_columns(df)
    
    df = filter_dates(df)
    
    df = filter_final_and_ceremony(df)
    df = filter_semifinal_participants(df)
    
    df = rename_games(df)
    
    df = clean_phone_numbers(df)
    df = clean_text_columns(df, reverse_name_mapping)
    
    df["Email"] = df["Email"].map(lambda x: repair_email_domain(x, domain_corrections))
    df = clean_comments(df)
    
    df = process_utm(df)
    df = determine_source(df)
    
    df = remove_duplicates(
        df,
        key_columns=["Игра", "Номер пакета", "Имя", "Email", "Телефон"],
        scenario="same_game_package"
    )
    df = remove_duplicates(
        df,
        key_columns=["Дата игры", "Имя", "Email", "Телефон"],
        scenario="same_date"
    )
    
    df = highlight_partial_duplicates(df)
    
    return df

# Сохранение в Excel
def save_to_excel(df: pd.DataFrame, output_file: str) -> None:    
    """
    Сохраняет обработанные данные в Excel с форматированием.
    Создаёт две вкладки: исходные данные (Sheet1) с удалением дубликатов и данные без изменений (No Duplicates).
    
    Args:
        df (pd.DataFrame): DataFrame для сохранения.
        output_file (str): Путь к выходному файлу.
    """
    logging.info(f"Сохранение результатов в файл: {output_file}")
    wb = Workbook()
    ws = wb.active
    ws.title = "registrations"

    blue_fill = PatternFill(start_color="00BBFF", end_color="00BBFF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

    # Проверяем наличие столбца _highlight_info
    has_highlight_info = "_highlight_info" in df.columns
    if has_highlight_info:
        highlight_info = df["_highlight_info"]
        df = df.drop(columns=["_highlight_info"])

    headers = df.columns.tolist()
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    date_game_col = headers.index("Дата игры") + 1 if "Дата игры" in headers else None
    phone_col = headers.index("Телефон") + 1 if "Телефон" in headers else None
    team_col = headers.index("Команда") + 1 if "Команда" in headers else None
    name_col = headers.index("Имя") + 1 if "Имя" in headers else None
    email_col = headers.index("Email") + 1 if "Email" in headers else None

    # Сохраняем форматирование для основной вкладки
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        highlight_cols = highlight_info.iloc[row_idx - 2] if has_highlight_info else []
        for col_idx, value in enumerate(row, start=1):
            if col_idx == date_game_col and pd.notna(value):
                value = value.date() if hasattr(value, "date") else value
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Проверяем, нужно ли подсветить ячейку оранжевым цветом
            if has_highlight_info and highlight_cols:
                if col_idx == team_col and "Команда" in highlight_cols:
                    cell.fill = orange_fill
                if col_idx == name_col and "Имя" in highlight_cols:
                    cell.fill = orange_fill
                elif col_idx == email_col and "Email" in highlight_cols:
                    cell.fill = orange_fill
                elif col_idx == phone_col and "Телефон" in highlight_cols:
                    # Проверяем, нужно ли применить жёлтую подсветку
                    if isinstance(value, str):
                        if not value.startswith(Config.PHONE_PREFIX):
                            cell.fill = blue_fill
                        elif len(value) != Config.EXPECTED_PHONE_LENGTH or (value.startswith("77") and value[2] == "9"):
                            cell.fill = yellow_fill
                        else:
                            cell.fill = orange_fill
                    else:
                        cell.fill = orange_fill
                continue

    # Настройка ширины столбцов для основной вкладки
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = max_length

    # Создание дополнительных вкладок для instagram, google, yandex
    source_tabs = {
        "instagram": df[df["source"] == "instagram"],
        "google": df[df["source"] == "google"],
        "yandex": df[df["source"] == "yandex"]
    }

    for tab_name, tab_df in source_tabs.items():
        if not tab_df.empty:
            logging.info(f"Создание вкладки '{tab_name}' с {len(tab_df)} строками")
            ws_tab = wb.create_sheet(title=tab_name)
            ws_tab.append(headers)
            for cell in ws_tab[1]:
                cell.font = Font(bold=True)
            ws_tab.freeze_panes = "A2"

            # Копируем форматирование для строк вкладки
            tab_highlight_info = highlight_info[tab_df.index] if has_highlight_info else pd.Series([[]] * len(tab_df), index=tab_df.index)
            for row_idx, row in enumerate(tab_df.itertuples(), start=2):
                df_idx = row.Index  # Получаем индекс строки из DataFrame
                highlight_cols = tab_highlight_info.loc[df_idx] if has_highlight_info else []
                for col_idx, value in enumerate(row[1:], start=1):  # Пропускаем первый элемент (Index)
                    if col_idx == date_game_col and pd.notna(value):
                        value = value.date() if hasattr(value, "date") else value
                    cell = ws_tab.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Проверяем, нужно ли подсветить ячейку оранжевым цветом
                    if has_highlight_info and highlight_cols:
                        if col_idx == team_col and "Команда" in highlight_cols:
                            cell.fill = orange_fill
                        if col_idx == name_col and "Имя" in highlight_cols:
                            cell.fill = orange_fill
                        elif col_idx == email_col and "Email" in highlight_cols:
                            cell.fill = orange_fill
                        elif col_idx == phone_col and "Телефон" in highlight_cols:
                            if isinstance(value, str):
                                if not value.startswith(Config.PHONE_PREFIX):
                                    cell.fill = blue_fill
                                elif len(value) != Config.EXPECTED_PHONE_LENGTH or (value.startswith("77") and value[2] == "9"):
                                    cell.fill = yellow_fill
                                else:
                                    cell.fill = orange_fill
                            else:
                                cell.fill = orange_fill
                        continue

            # Настройка ширины столбцов для вкладки
            for column in ws_tab.columns:
                max_length = max(len(str(cell.value or "")) for cell in column) + 2
                ws_tab.column_dimensions[column[0].column_letter].width = max_length
        else:
            logging.info(f"Вкладка '{tab_name}' не создана, так как строк с source='{tab_name}' не найдено")

    # Создание вкладки unique с уникальными комбинациями Имя, Email, Телефон (игнорируя Команда при сравнении)
    unique_columns = ["Команда", "Имя", "Email", "Телефон"]
    unique_key_columns = ["Имя", "Email", "Телефон"]  # Столбцы для определения уникальности
    if all(col in df.columns for col in unique_columns):
        logging.info("Создание вкладки 'unique' с уникальными комбинациями (игнорируя 'Команда')")
        # Удаляем дубликаты по Имя, Email, Телефон, сохраняя первую строку
        unique_df = df[unique_columns].drop_duplicates(subset=unique_key_columns, keep="first").reset_index(drop=True)
        unique_highlight_info = pd.Series([[]] * len(unique_df), index=unique_df.index)

        # Копируем форматирование для уникальных строк
        if has_highlight_info:
            unique_highlight_info = pd.Series([[]] * len(unique_df), index=unique_df.index)
            for idx, row in unique_df.iterrows():
                # Находим первую строку в исходном DataFrame с такой комбинацией Имя, Email, Телефон
                mask = (df[unique_key_columns] == row[unique_key_columns]).all(axis=1)
                if mask.any():
                    orig_idx = df[mask].index[0]
                    unique_highlight_info.iloc[idx] = highlight_info.iloc[orig_idx]

        ws_unique = wb.create_sheet(title="unique")
        ws_unique.append(unique_columns)
        for cell in ws_unique[1]:
            cell.font = Font(bold=True)
        ws_unique.freeze_panes = "A2"

        # Копируем форматирование для строк вкладки unique
        unique_phone_col = unique_columns.index("Телефон") + 1

        for row_idx, row in enumerate(unique_df.itertuples(index=False), start=2):
            highlight_cols = unique_highlight_info.iloc[row_idx - 2] if has_highlight_info else []
            for col_idx, value in enumerate(row, start=1):
                cell = ws_unique.cell(row=row_idx, column=col_idx, value=value)
                
                # Копируем форматирование телефонов
                if col_idx == unique_phone_col and isinstance(value, str):
                    if not value.startswith(Config.PHONE_PREFIX):
                        cell.fill = blue_fill
                    elif len(value) != Config.EXPECTED_PHONE_LENGTH or (value.startswith("77") and value[2] == "9"):
                        cell.fill = yellow_fill

        # Настройка ширины столбцов для вкладки unique
        for column in ws_unique.columns:
            max_length = max(len(str(cell.value or "")) for cell in column) + 2
            ws_unique.column_dimensions[column[0].column_letter].width = max_length
        logging.info(f"Вкладка 'unique' создана с {len(unique_df)} уникальными строками (игнорируя 'Команда')")
    else:
        logging.warning("Не удалось создать вкладку 'unique', так как отсутствуют необходимые столбцы")

    try:
        wb.save(output_file)
        logging.info(f"Файл успешно сохранён: {output_file}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла: {type(e).__name__} - {e}")
        raise

# Основная функция
def main():
    parser = argparse.ArgumentParser(description="Обработка данных из Excel-файла")
    parser.add_argument("--input", default=r"C:\Users\Artem\Downloads\grid-export (1).xlsx", help="Путь к входному файлу")
    parser.add_argument("--output", default=r"C:\Users\Artem\Downloads\potter.xlsx", help="Путь к выходному файлу")
    parser.add_argument("--log-level", default="INFO", help="Уровень логирования")
    args = parser.parse_args()

    setup_logging(args.log_level)
    error_count = 0

    try:
        name_variants, domain_corrections, reverse_name_mapping = load_dictionaries()
    except Exception as e:
        logging.error(f"Ошибка при загрузке словарей: {type(e).__name__} - {e}")
        error_count += 1
        return error_count

    try:
        raw_data = pd.read_excel(args.input)
        logging.info(f"Файл успешно прочитан: {args.input}, строк: {len(raw_data)}")
    except Exception as e:
        logging.error(f"Ошибка при чтении файла {args.input}: {type(e).__name__} - {e}")
        error_count += 1
        return error_count

    try:
        cleaned_data = clean_data(raw_data, reverse_name_mapping, domain_corrections)
        logging.info(f"Данные успешно обработаны, строк после очистки: {len(cleaned_data)}")
    except Exception as e:
        logging.error(f"Ошибка при очистке данных: {type(e).__name__} - {e}")
        error_count += 1
        return error_count

    try:
        save_to_excel(cleaned_data, args.output)
    except Exception as e:
        logging.error(f"Ошибка при сохранении результатов: {type(e).__name__} - {e}")
        error_count += 1
        return error_count

    if error_count > 0:
        logging.error(f"Обработка завершена с ошибками: {error_count}")
    else:
        logging.info("Обработка успешно завершена")
    return error_count

if __name__ == "__main__":
    main()
